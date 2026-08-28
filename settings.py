from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, send_file
import csv
import io
from flask_login import login_required, current_user
from flask_login import login_required, current_user
from models import db, Site, Category, User, Location, SubCategory, WorkOrderStatus, ProjectCode, Tasklist, TasklistProcedure, Checklist, ChecklistParameterTemplate, ChillerFaultCode, Team, HelpdeskModule, HelpdeskLocation, Part, StockLevel, Customer, LogsheetTemplate, LogsheetTemplateParameter

settings_bp = Blueprint('settings', __name__, url_prefix='/settings')

@settings_bp.before_request
def restrict_to_admin():
    if not current_user.is_authenticated or current_user.role != 'Admin':
        from flask import flash, redirect, url_for
        flash('Access denied. Administrator privileges required.', 'danger')
        return redirect(url_for('dashboard'))

@settings_bp.route('/')
@login_required
def index():
    if current_user.role != 'Admin':
        flash('You do not have permission to access Settings.', 'error')
        return redirect(url_for('dashboard'))
        
    if current_user.site_id:
        sites = Site.query.filter_by(id=current_user.site_id).all()
        locations = Location.query.filter_by(site_id=current_user.site_id).all()
        users = User.query.filter_by(site_id=current_user.site_id).all()
        teams = Team.query.filter_by(site_id=current_user.site_id).all()
        hd_modules = HelpdeskModule.query.filter((HelpdeskModule.site_id == current_user.site_id) | (HelpdeskModule.site_id == None)).all()
        hd_locations = HelpdeskLocation.query.filter((HelpdeskLocation.site_id == current_user.site_id) | (HelpdeskLocation.site_id == None)).all()
    else:
        sites = Site.query.all()
        locations = Location.query.all()
        users = User.query.all()
        teams = Team.query.all()
        hd_modules = HelpdeskModule.query.all()
        hd_locations = HelpdeskLocation.query.all()
        
    categories = Category.query.all()
    subcategories = SubCategory.query.all()
    statuses = WorkOrderStatus.query.all()
    if current_user.site_id:
        user_site = Site.query.get(current_user.site_id)
        if user_site and user_site.project_code:
            project_codes = ProjectCode.query.filter_by(code=user_site.project_code).all()
            tasklists = Tasklist.query.filter(db.or_(Tasklist.project_code == user_site.project_code, Tasklist.project_code == None, Tasklist.project_code == '')).order_by(Tasklist.name).all()
            checklists = Checklist.query.filter(db.or_(Checklist.project_code == user_site.project_code, Checklist.project_code == None, Checklist.project_code == '')).order_by(Checklist.name).all()
        else:
            project_codes = []
            tasklists = Tasklist.query.filter(db.or_(Tasklist.project_code == None, Tasklist.project_code == '')).order_by(Tasklist.name).all()
            checklists = Checklist.query.filter(db.or_(Checklist.project_code == None, Checklist.project_code == '')).order_by(Checklist.name).all()
    else:
        project_codes = ProjectCode.query.all()
        tasklists = Tasklist.query.order_by(Tasklist.name).all()
        checklists = Checklist.query.order_by(Checklist.name).all()
        
    for tl in tasklists:
        tl.project_code_group = tl.project_code if tl.project_code else ""
        
    for cl in checklists:
        cl.project_code_group = cl.project_code if cl.project_code else ""
    
    chiller_faults_raw = ChillerFaultCode.query.all()
    chiller_faults = {}
    for cf in chiller_faults_raw:
        if cf.chiller_type not in chiller_faults:
            chiller_faults[cf.chiller_type] = {}
        if cf.category not in chiller_faults[cf.chiller_type]:
            chiller_faults[cf.chiller_type][cf.category] = []
        chiller_faults[cf.chiller_type][cf.category].append(cf)
        
    parts = Part.query.all()
    customers = Customer.query.all()
    logsheet_templates = LogsheetTemplate.query.order_by(LogsheetTemplate.name).all()
    
    return render_template('settings/index.html', sites=sites, categories=categories, users=users, locations=locations, subcategories=subcategories, statuses=statuses, project_codes=project_codes, tasklists=tasklists, checklists=checklists, chiller_faults=chiller_faults, teams=teams, hd_modules=hd_modules, hd_locations=hd_locations, parts=parts, customers=customers, logsheet_templates=logsheet_templates)

@settings_bp.route('/site/add', methods=['POST'])
@login_required
def add_site():
    if current_user.role != 'Admin' or current_user.site_id is not None:
        flash('You do not have permission to add sites.', 'error')
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    address = request.form.get('address')
    timezone = request.form.get('timezone', 'UTC')
    api_site_id = request.form.get('api_site_id') or None
    project_code = request.form.get('project_code')
    
    if not project_code or project_code.strip() == '':
        project_code = None
    
    if name:
        new_site = Site(name=name, address=address, timezone=timezone, api_site_id=api_site_id, project_code=project_code)
        db.session.add(new_site)
        db.session.commit()
        flash('Site added successfully.', 'success')
        
    return redirect(url_for('settings.index'))

@settings_bp.route('/site/<int:id>/edit', methods=['POST'])
@login_required
def edit_site(id):
    if current_user.role != 'Admin' or (current_user.site_id is not None and current_user.site_id != id):
        flash('You do not have permission to edit this site.', 'error')
        return redirect(url_for('dashboard'))
        
    site = Site.query.get_or_404(id)
    site.name = request.form.get('name', site.name)
    site.address = request.form.get('address', site.address)
    site.timezone = request.form.get('timezone', site.timezone)
    site.api_site_id = request.form.get('api_site_id') or None
    
    project_code = request.form.get('project_code')
    if not project_code or project_code.strip() == '':
        site.project_code = None
    else:
        site.project_code = project_code
        
    db.session.commit()
    flash('Site updated successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#sites')

@settings_bp.route('/site/<int:id>/delete', methods=['POST'])
@login_required
def delete_site(id):
    if current_user.role != 'Admin' or (current_user.site_id is not None and current_user.site_id != id):
        flash('You do not have permission to delete this site.', 'error')
        return redirect(url_for('dashboard'))
        
    site = Site.query.get_or_404(id)
    
    # Cascading Nuke: Delete all child Locations and Assets first to prevent SQLite Integrity Errors
    try:
        # Delete Assets First (Leaf Nodes)
        from models import Asset
        assets_to_delete = Asset.query.filter_by(site_id=site.id).all()
        for ast in assets_to_delete:
            db.session.delete(ast)
            
        # Delete Locations Next (Middle Nodes)
        locations_to_delete = Location.query.filter_by(site_id=site.id).all()
        for loc in locations_to_delete:
            db.session.delete(loc)
            
        db.session.delete(site)
        db.session.commit()
        flash('Site dan SELURUH Asset serta Lokasi di dalamnya telah hangus dihapus secara permanen!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus Site akibat struktur data kompleks yang menyangkut: {str(e)}', 'danger')
        
    return redirect(url_for('settings.index') + '#sites')

@settings_bp.route('/site/<int:id>/safe_ranges/save', methods=['POST'])
@login_required
def save_safe_ranges(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    from models import SiteSafeRange
    
    try:
        # Clear existing ranges for this site
        SiteSafeRange.query.filter_by(site_id=id).delete()
        
        keys = request.form.getlist('param_key[]')
        mins = request.form.getlist('min_val[]')
        maxs = request.form.getlist('max_val[]')
        
        count = 0
        for k, m_min, m_max in zip(keys, mins, maxs):
            k = str(k).strip()
            if not k:
                continue
                
            m_min_val = None
            if m_min and str(m_min).strip() != '':
                m_min_val = float(str(m_min).strip())
                
            m_max_val = None
            if m_max and str(m_max).strip() != '':
                m_max_val = float(str(m_max).strip())
                
            if m_min_val is None and m_max_val is None:
                continue
                
            new_range = SiteSafeRange(site_id=id, parameter_key=k, min_value=m_min_val, max_value=m_max_val)
            db.session.add(new_range)
            count += 1
            
        db.session.commit()
        flash(f'Successfully saved {count} safe ranges for the site.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving safe ranges: {str(e)}', 'danger')
        
    return redirect(url_for('settings.index') + '#sites')

@settings_bp.route('/site/<int:id>/safe_ranges/export')
@login_required
def export_safe_ranges(id):
    from models import Site
    import pandas as pd
    import io
    from flask import send_file
    
    site = Site.query.get_or_404(id)
    data = []
    for r in site.safe_ranges:
        data.append({
            'Parameter Key': r.parameter_key,
            'Min Value': r.min_value if r.min_value is not None else '',
            'Max Value': r.max_value if r.max_value is not None else ''
        })
    df = pd.DataFrame(data, columns=['Parameter Key', 'Min Value', 'Max Value'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    filename = f'safe_ranges_{site.name.replace(" ", "_")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@settings_bp.route('/location/add', methods=['POST'])
@login_required
def add_location():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    description = request.form.get('description')
    site_id = request.form.get('site_id')
    category_id = request.form.get('category_id') or None
    
    if name and site_id:
        new_loc = Location(name=name, description=description, site_id=site_id, category_id=category_id)
        db.session.add(new_loc)
        db.session.commit()
        flash('Location added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#sites')

@settings_bp.route('/location/<int:id>/delete', methods=['POST'])
@login_required
def delete_location(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    loc = Location.query.get_or_404(id)
    try:
        db.session.delete(loc)
        db.session.commit()
        flash('Location deleted successfully.', 'success')
    except Exception:
        db.session.rollback()
        flash('Gagal menghapus Lokasi karena masih ada Aset yang terhubung dengannya!', 'danger')
    return redirect(url_for('settings.index') + '#sites')

@settings_bp.route('/category/add', methods=['POST'])
@login_required
def add_category():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    type = request.form.get('type', 'Asset')
    
    if name:
        new_cat = Category(name=name, type=type)
        db.session.add(new_cat)
        db.session.commit()
        flash('Category added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#categories')

@settings_bp.route('/category/<int:id>/delete', methods=['POST'])
@login_required
def delete_category(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    cat = Category.query.get_or_404(id)
    try:
        db.session.delete(cat)
        db.session.commit()
        flash('Category deleted successfully.', 'success')
    except Exception:
        db.session.rollback()
        flash('Gagal menghapus Kategori karena masih ada Aset yang menggunakannya!', 'danger')
    return redirect(url_for('settings.index') + '#categories')

@settings_bp.route('/subcategory/add', methods=['POST'])
@login_required
def add_subcategory():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    category_id = request.form.get('category_id')
    
    if name and category_id:
        new_subcat = SubCategory(name=name, category_id=category_id)
        db.session.add(new_subcat)
        db.session.commit()
        flash('Subcategory added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#categories')

@settings_bp.route('/subcategory/<int:id>/delete', methods=['POST'])
@login_required
def delete_subcategory(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    subcat = SubCategory.query.get_or_404(id)
    db.session.delete(subcat)
    db.session.commit()
    flash('Subcategory deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#categories')

@settings_bp.route('/wo_status/add', methods=['POST'])
@login_required
def add_wo_status():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    description = request.form.get('description')
    state_type = request.form.get('state_type', 'Open')
    
    if name:
        new_status = WorkOrderStatus(name=name, description=description, state_type=state_type)
        db.session.add(new_status)
        db.session.commit()
        flash('Work Order Status added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#wo-statuses')

@settings_bp.route('/wo_status/<int:id>/edit', methods=['POST'])
@login_required
def edit_wo_status(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    status = WorkOrderStatus.query.get_or_404(id)
    
    name = request.form.get('name')
    description = request.form.get('description')
    state_type = request.form.get('state_type')
    
    if name:
        status.name = name
        status.description = description
        if state_type:
            status.state_type = state_type
            
        db.session.commit()
        flash('Work Order Status updated successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#wo-statuses')

@settings_bp.route('/wo_status/<int:id>/delete', methods=['POST'])
@login_required
def delete_wo_status(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    status = WorkOrderStatus.query.get_or_404(id)
    db.session.delete(status)
    db.session.commit()
    flash('Work Order Status deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#statuses')

@settings_bp.route('/project_code/add', methods=['POST'])
@login_required
def add_project_code():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    code = request.form.get('code')
    description = request.form.get('description')
    
    if code:
        new_proj = ProjectCode(code=code, description=description)
        db.session.add(new_proj)
        db.session.commit()
        flash('Project Code added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#project-codes')

@settings_bp.route('/project_code/<int:id>/delete', methods=['POST'])
@login_required
def delete_project_code(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    proj = ProjectCode.query.get_or_404(id)
    db.session.delete(proj)
    db.session.commit()
    flash('Project Code deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#project-codes')


@settings_bp.route('/team/add', methods=['POST'])
@login_required
def add_team():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    description = request.form.get('description')
    site_id = request.form.get('site_id')
    user_ids = request.form.getlist('user_ids')
    
    if name and site_id:
        new_team = Team(name=name, description=description, site_id=site_id)
        db.session.add(new_team)
        db.session.commit() # commit team to get ID
        
        if user_ids:
            users_to_assign = User.query.filter(User.id.in_(user_ids)).all()
            for u in users_to_assign:
                u.team_id = new_team.id
                u.site_id = new_team.site_id
            db.session.commit()
            
        flash('Team added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#teams')

@settings_bp.route('/team/<int:id>/edit', methods=['POST'])
@login_required
def edit_team(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    team = Team.query.get_or_404(id)
    team.name = request.form.get('name', team.name)
    team.description = request.form.get('description')
    
    site_id = request.form.get('site_id')
    user_ids = request.form.getlist('user_ids')
    
    if site_id:
        team.site_id = site_id
        
    # Unassign users currently in team but not in new list
    for u in team.users:
        if str(u.id) not in user_ids:
            u.team_id = None
            
    # Assign new users
    if user_ids:
        users_to_assign = User.query.filter(User.id.in_(user_ids)).all()
        for u in users_to_assign:
            u.team_id = team.id
            u.site_id = team.site_id
            
    db.session.commit()
    flash('Team updated successfully.', 'success')
    return redirect(url_for('settings.index') + '#teams')

@settings_bp.route('/team/<int:id>/delete', methods=['POST'])
@login_required
def delete_team(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    team = Team.query.get_or_404(id)
    
    # Optional: Unassign users from this team
    for u in team.users:
        u.team_id = None
        
    db.session.delete(team)
    db.session.commit()
    flash('Team deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#teams')

@settings_bp.route('/user/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    nrp = request.form.get('nrp')
    name = request.form.get('name')
    password = request.form.get('password')
    role = request.form.get('role', 'Technician')
    site_id = request.form.get('site_id')
    team_id = request.form.get('team_id')
    phone_number = request.form.get('phone_number')
    
    if not team_id or team_id == '0' or team_id == '':
        team_id = None
        
    if team_id:
        team_obj = db.session.get(Team, team_id)
        if team_obj:
            site_id = team_obj.site_id
            
    if not site_id or site_id == '0' or site_id == '':
        site_id = None
    
    if nrp and name and password:
        if User.query.filter_by(nrp=nrp).first():
            flash('NRP already exists.', 'error')
            return redirect(url_for('settings.index') + '#users')
            
        new_user = User(nrp=nrp, name=name, role=role, site_id=site_id, team_id=team_id, phone_number=phone_number, is_approved=True)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        flash(f'User {name} added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#users')

@settings_bp.route('/user/<int:id>/edit', methods=['POST'])
@login_required
def edit_user(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    user = User.query.get_or_404(id)
    name = request.form.get('name')
    nrp = request.form.get('nrp')
    role = request.form.get('role')
    site_id = request.form.get('site_id')
    team_id = request.form.get('team_id')
    phone_number = request.form.get('phone_number')
    password = request.form.get('password') # Optional
    is_approved = request.form.get('is_approved') == 'on'
    
    if not team_id or team_id == '0' or team_id == '':
        team_id = None
        
    if team_id:
        team_obj = db.session.get(Team, team_id)
        if team_obj:
            site_id = team_obj.site_id
            
    if not site_id or site_id == '0' or site_id == '':
        site_id = None
        
    if name and nrp and role:
        # Check if the new NRP conflicts with another user
        existing = User.query.filter(User.nrp == nrp, User.id != id).first()
        if existing:
            flash('Another user with that NRP already exists.', 'error')
        else:
            user.name = name
            user.nrp = nrp
            user.role = role
            user.site_id = site_id
            user.team_id = team_id
            user.phone_number = phone_number
            user.is_approved = is_approved
            if password:
                user.set_password(password)
            db.session.commit()
            flash(f'User {name} updated successfully.', 'success')
            
    return redirect(url_for('settings.index') + '#users')

@settings_bp.route('/user/<int:id>/delete', methods=['POST'])
@login_required
def delete_user(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('Cannot delete your own account!', 'error')
        return redirect(url_for('settings.index') + '#users')
        
    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#users')

@settings_bp.route('/users/bulk_action', methods=['POST'])
@login_required
def bulk_action_users():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    action = request.form.get('action')
    user_ids = request.form.getlist('user_ids')
    
    if not user_ids:
        flash('No users selected.', 'warning')
        return redirect(url_for('settings.index') + '#users')
        
    if action == 'approve':
        for uid in user_ids:
            u = db.session.get(User, uid)
            if u and not u.is_approved:
                u.is_approved = True
        db.session.commit()
        flash(f'Successfully approved {len(user_ids)} users.', 'success')
    elif action == 'delete':
        for uid in user_ids:
            u = db.session.get(User, uid)
            if u and u.id != current_user.id:
                db.session.delete(u)
        db.session.commit()
        flash(f'Successfully deleted {len(user_ids)} users.', 'success')
        
    return redirect(url_for('settings.index') + '#users')

@settings_bp.route('/user/export', methods=['GET'])
@login_required
def export_users():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    import pandas as pd
    
    users = User.query.all()
    data = []
    for u in users:
        site_name = u.site.name if u.site else 'Kantor'
        data.append({
            'NRP': u.nrp,
            'Name': u.name,
            'Role': u.role,
            'Site': site_name,
            'Password': ''
        })
        
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Users')
        
    output.seek(0)
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": "attachment; filename=users.xlsx"}
    )

@settings_bp.route('/user/template', methods=['GET'])
@login_required
def download_user_template():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    import pandas as pd
    
    # Create template with headers and 1 sample row
    data = [{
        'NRP': '1234567890',
        'Name': 'John Doe',
        'Role': 'Technician',
        'Site': 'Kantor',
        'Password': 'password123'
    }]
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Users')
        
        # Adjust column widths
        worksheet = writer.sheets['Users']
        worksheet.set_column('A:E', 20)
        
    output.seek(0)
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": "attachment; filename=users_template.xlsx"}
    )

@settings_bp.route('/user/import', methods=['POST'])
@login_required
def import_users():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    import pandas as pd
    
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('Invalid file format. Please upload an Excel (.xlsx) file.', 'error')
        return redirect(url_for('settings.index') + '#users')
        
    try:
        df = pd.read_excel(file)
        # Ensure we don't process NaN as valid data
        df = df.fillna('')
        
        counter = 0
        for index, row in df.iterrows():
            nrp = str(row.get('NRP', '')).strip()
            name = str(row.get('Name', '')).strip()
            role = str(row.get('Role', 'Technician')).strip()
            site_name = str(row.get('Site', '')).strip()
            password = str(row.get('Password', '')).strip()
            
            # Since pandas might convert large numbers to float representations, clean up NRP
            if nrp.endswith('.0'):
                nrp = nrp[:-2]
                
            if not nrp or not name:
                continue
                
            site_id = None
            if site_name and site_name.lower() != 'kantor':
                site_obj = Site.query.filter(Site.name.ilike(site_name)).first()
                if site_obj:
                    site_id = site_obj.id
                
            existing_user = User.query.filter_by(nrp=nrp).first()
            if not existing_user:
                new_user = User(nrp=nrp, name=name, role=role, site_id=site_id)
                if not password:
                    password = nrp  # Default to NRP if no password provided
                new_user.set_password(password)
                db.session.add(new_user)
                counter += 1
        db.session.commit()
        flash(f'Users imported successfully. Added {counter} new users.', 'success')
    except Exception as e:
        flash(f'Error importing users: {str(e)}', 'error')
        
    return redirect(url_for('settings.index') + '#users')

@settings_bp.route('/tasklist/add', methods=['POST'])
@login_required
def add_tasklist():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    description = request.form.get('description')
    project_code = request.form.get('project_code')
    
    if not project_code or project_code.strip() == '':
        project_code = None
    
    if name:
        new_tl = Tasklist(name=name, description=description, project_code=project_code)
        db.session.add(new_tl)
        db.session.flush() # flush to get the new_tl.id!
        
        procedures_json = request.form.get('procedures_json')
        if procedures_json:
            import json
            try:
                proc_data_list = json.loads(procedures_json)
                for idx, proc_data in enumerate(proc_data_list):
                    proc_name = proc_data.get('name', '').strip()
                    req_attach = proc_data.get('requires_attachment', False)
                    min_photos = proc_data.get('min_photos', 0)
                    if proc_name:
                        new_p = TasklistProcedure(
                            tasklist_id=new_tl.id,
                            name=proc_name,
                            requires_attachment=req_attach,
                            min_photos=min_photos,
                            position=idx
                        )
                        db.session.add(new_p)
            except Exception as e:
                flash(f'Error processing procedures: {str(e)}', 'error')
                
        db.session.commit()
        flash('Tasklist added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#tasklists')

@settings_bp.route('/tasklist/<int:id>/delete', methods=['POST'])
@login_required
def delete_tasklist(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    tl = Tasklist.query.get_or_404(id)
    db.session.delete(tl)
    db.session.commit()
    flash('Tasklist deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#tasklists')

@settings_bp.route('/tasklist/<int:id>/edit', methods=['POST'])
@login_required
def edit_tasklist(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    tl = Tasklist.query.get_or_404(id)
    name = request.form.get('name')
    description = request.form.get('description')
    project_code = request.form.get('project_code')
    procedures_json = request.form.get('procedures_json')
    
    if not project_code or project_code.strip() == '':
        project_code = None
        
    if name:
        tl.name = name
        tl.description = description
        tl.project_code = project_code
        
        # Handle embedded procedures if provided
        if procedures_json:
            import json
            try:
                proc_data_list = json.loads(procedures_json)
                current_procs_dict = {p.id: p for p in tl.procedures}
                kept_ids = set()
                
                for idx, proc_data in enumerate(proc_data_list):
                    proc_id = proc_data.get('id')
                    proc_name = proc_data.get('name', '').strip()
                    req_attach = proc_data.get('requires_attachment', False)
                    
                    if not proc_name:
                        continue
                        
                    if proc_id and str(proc_id).startswith('new_'):
                        # New procedure
                        new_p = TasklistProcedure(
                            tasklist_id=tl.id,
                            name=proc_name,
                            requires_attachment=req_attach,
                            min_photos=proc_data.get('min_photos', 0),
                            position=idx
                        )
                        db.session.add(new_p)
                    elif proc_id and int(proc_id) in current_procs_dict:
                        # Update existing
                        pid = int(proc_id)
                        p = current_procs_dict[pid]
                        p.name = proc_name
                        p.requires_attachment = req_attach
                        p.min_photos = proc_data.get('min_photos', 0)
                        p.position = idx
                        kept_ids.add(pid)
                        
                # Delete removed procedures
                for pid, p in current_procs_dict.items():
                    if pid not in kept_ids:
                        db.session.delete(p)
            except Exception as e:
                flash(f'Error processing procedures: {str(e)}', 'error')

        db.session.commit()
        flash('Tasklist updated successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#tasklists')

@settings_bp.route('/tasklist_procedure/add', methods=['POST'])
@login_required
def add_tasklist_procedure():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    tasklist_id = request.form.get('tasklist_id')
    name = request.form.get('name')
    requires_attachment = 'requires_attachment' in request.form
    
    if tasklist_id and name:
        new_proc = TasklistProcedure(tasklist_id=tasklist_id, name=name, requires_attachment=requires_attachment)
        db.session.add(new_proc)
        db.session.commit()
        flash('Tasklist Procedure added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#tasklists')

@settings_bp.route('/api/tasklist/<int:id>/procedures')
@login_required
def api_tasklist_procedures(id):
    if current_user.role != 'Admin':
        return {'procedures': []}
    tl = Tasklist.query.get_or_404(id)
    out = []
    for p in tl.procedures:
        out.append({'id': p.id, 'name': p.name, 'requires_attachment': p.requires_attachment, 'min_photos': p.min_photos})
    return {'procedures': out}

@settings_bp.route('/tasklist_procedure/<int:id>/delete', methods=['POST'])
@login_required
def delete_tasklist_procedure(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    proc = TasklistProcedure.query.get_or_404(id)
    db.session.delete(proc)
    db.session.commit()
    flash('Tasklist Procedure deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#tasklists')

@settings_bp.route('/checklist/add', methods=['POST'])
@login_required
def add_checklist():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    name = request.form.get('name')
    project_code = request.form.get('project_code')
    
    if not project_code or project_code.strip() == '':
        project_code = None
    
    if name:
        new_cl = Checklist(name=name, project_code=project_code)
        db.session.add(new_cl)
        db.session.commit()
        flash('Checklist template added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#checklists')

@settings_bp.route('/checklist/<int:id>/delete', methods=['POST'])
@login_required
def delete_checklist(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    cl = Checklist.query.get_or_404(id)
    # Using raw execute or cascade will delete parameters. For simplicity:
    for p in cl.parameters:
        db.session.delete(p)
    db.session.delete(cl)
    db.session.commit()
    flash('Checklist template deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#checklists')

@settings_bp.route('/checklist/<int:id>/edit', methods=['POST'])
@login_required
def edit_checklist(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    cl = Checklist.query.get_or_404(id)
    name = request.form.get('name')
    project_code = request.form.get('project_code')
    parameters_json = request.form.get('parameters_json')
    
    if not project_code or project_code.strip() == '':
        project_code = None
        
    if name:
        existing = Checklist.query.filter(Checklist.name == name, Checklist.id != id).first()
        if existing:
            flash(f'A checklist named {name} already exists.', 'error')
        else:
            cl.name = name
            cl.project_code = project_code
            
            # Handle embedded parameters if provided
            if parameters_json:
                import json
                try:
                    param_data_list = json.loads(parameters_json)
                    current_params_dict = {p.id: p for p in cl.parameters}
                    kept_ids = set()
                    
                    for idx, pdata in enumerate(param_data_list):
                        pid_val = pdata.get('id')
                        param_name = pdata.get('parameter', '').strip()
                        standard = pdata.get('standard', '').strip()
                        
                        if not param_name:
                            continue
                            
                        if pid_val and str(pid_val).startswith('new_'):
                            # New parameter
                            new_p = ChecklistParameterTemplate(
                                checklist_id=cl.id,
                                parameter=param_name,
                                standard=standard,
                                position=idx
                            )
                            db.session.add(new_p)
                        elif pid_val and int(pid_val) in current_params_dict:
                            # Update existing
                            int_id = int(pid_val)
                            p = current_params_dict[int_id]
                            p.parameter = param_name
                            p.standard = standard
                            p.position = idx
                            kept_ids.add(int_id)
                            
                    # Delete removed parameters
                    for pid, p in current_params_dict.items():
                        if pid not in kept_ids:
                            db.session.delete(p)
                except Exception as e:
                    flash(f'Error processing parameters: {str(e)}', 'error')
            
            db.session.commit()
            flash('Checklist updated successfully.', 'success')
            
    return redirect(url_for('settings.index') + '#checklists')

@settings_bp.route('/checklist_parameter/add', methods=['POST'])
@login_required
def add_checklist_parameter():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    checklist_id = request.form.get('checklist_id')
    parameter = request.form.get('parameter')
    standard = request.form.get('standard')
    
    if checklist_id and parameter and standard:
        new_param = ChecklistParameterTemplate(checklist_id=checklist_id, parameter=parameter, standard=standard)
        db.session.add(new_param)
        db.session.commit()
        flash('Checklist Parameter added successfully.', 'success')
        
    return redirect(url_for('settings.index') + '#checklists')

@settings_bp.route('/api/checklist/<int:id>/parameters')
@login_required
def api_checklist_parameters(id):
    if current_user.role != 'Admin':
        return {'parameters': []}
    cl = Checklist.query.get_or_404(id)
    out = []
    for p in cl.parameters:
        out.append({'id': p.id, 'parameter': p.parameter, 'standard': p.standard})
    return {'parameters': out}

@settings_bp.route('/checklist_parameter/<int:id>/delete', methods=['POST'])
@login_required
def delete_checklist_parameter(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    param = ChecklistParameterTemplate.query.get_or_404(id)
    db.session.delete(param)
    db.session.commit()
    flash('Checklist Parameter deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#checklists')

import pandas as pd

@settings_bp.route('/checklist/template', methods=['GET'])
@login_required
def download_checklist_template():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    df = pd.DataFrame(columns=['Checklist Name', 'Project Code', 'Parameter', 'Standard'])
    df.loc[0] = ['Contoh Checklist AC', 'PRJ-001', 'Suhu Ruangan', '18-24 Celcius']
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(
        output,
        download_name="checklist_template.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@settings_bp.route('/checklist/export', methods=['GET'])
@login_required
def export_checklists():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    data = []
    checklists = Checklist.query.all()
    for cl in checklists:
        if cl.parameters.count() == 0:
            data.append({
                'Checklist Name': cl.name,
                'Project Code': cl.project_code or '',
                'Parameter': '',
                'Standard': ''
            })
        for p in cl.parameters:
            data.append({
                'Checklist Name': cl.name,
                'Project Code': cl.project_code or '',
                'Parameter': p.parameter,
                'Standard': p.standard
            })
            
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(
        output,
        download_name="checklists.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@settings_bp.route('/checklist/import', methods=['POST'])
@login_required
def import_checklists():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'error')
        return redirect(url_for('settings.index'))
        
    try:
        df = pd.read_excel(file)
        
        counter = 0
        for index, row in df.iterrows():
            cl_name = str(row.get('Checklist Name', '')).strip()
            if cl_name == 'nan': cl_name = ''
            
            p_code = str(row.get('Project Code', '')).strip()
            if p_code == 'nan': p_code = ''
                
            param = str(row.get('Parameter', '')).strip()
            if param == 'nan': param = ''
                
            st = str(row.get('Standard', '')).strip()
            if st == 'nan': st = ''
            
            if not cl_name:
                continue
                
            cl = Checklist.query.filter_by(name=cl_name).first()
            if not cl:
                cl = Checklist(name=cl_name, project_code=p_code if p_code else None)
                db.session.add(cl)
                db.session.flush()
            else:
                if p_code and cl.project_code != p_code:
                    cl.project_code = p_code
                
            if param and st:
                existing = ChecklistParameterTemplate.query.filter_by(checklist_id=cl.id, parameter=param).first()
                if not existing:
                    new_param = ChecklistParameterTemplate(checklist_id=cl.id, parameter=param, standard=st)
                    db.session.add(new_param)
                    counter += 1
        db.session.commit()
        flash(f'Checklists imported smoothly. Added {counter} new parameters.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing Excel: {str(e)}', 'error')
        
    return redirect(url_for('settings.index'))

@settings_bp.route('/tasklist/template', methods=['GET'])
@login_required
def download_tasklist_template():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    df = pd.DataFrame(columns=['Name', 'Project Code', 'Task/Step Name', 'Task/Attachment?', 'Task/Duration (Minutes)'])
    df.loc[0] = ['Contoh PM AC Split', 'PRJ-001', 'Bersihkan Filter Udara', 'TRUE', '15']
    df.loc[1] = ['Contoh PM AC Split', 'PRJ-001', 'Cek Tekanan Freon', 'FALSE', '10']
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(
        output,
        download_name="tasklist_template.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@settings_bp.route('/tasklist/export', methods=['GET'])
@login_required
def export_tasklists():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    data = []
    tasklists = Tasklist.query.all()
    for tl in tasklists:
        if tl.procedures.count() == 0:
            data.append({
                'Name': tl.name,
                'Project Code': tl.project_code or '',
                'Task/Step Name': '',
                'Task/Attachment?': '',
                'Task/Duration (Minutes)': ''
            })
        for p in tl.procedures:
            req_att = 'TRUE' if p.requires_attachment else 'FALSE'
            data.append({
                'Name': tl.name,
                'Project Code': tl.project_code or '',
                'Task/Step Name': p.name,
                'Task/Attachment?': req_att,
                'Task/Duration (Minutes)': p.estimated_minutes
            })
            
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(
        output,
        download_name="tasklists.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@settings_bp.route('/tasklist/import', methods=['POST'])
@login_required
def import_tasklists():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'error')
        return redirect(url_for('settings.index'))
        
    try:
        df = pd.read_excel(file)
            
        counter = 0
        last_tl_name = None
        for index, row in df.iterrows():
            tl_name = str(row.get('Name', '')).strip()
            if tl_name == 'nan': tl_name = ''
                
            if not tl_name and last_tl_name:
                tl_name = last_tl_name
            else:
                last_tl_name = tl_name
                
            p_code = str(row.get('Project Code', '')).strip()
            if p_code == 'nan': p_code = ''
                
            proc = str(row.get('Task/Step Name', '')).strip()
            if proc == 'nan': proc = ''
            
            req_raw = str(row.get('Task/Attachment?', '')).strip().upper()
            req_att = True if req_raw == 'TRUE' else False
            
            dur_raw = str(row.get('Task/Duration (Minutes)', '0')).strip()
            if dur_raw == 'nan': dur_raw = '0'
            try:
                dur_raw = dur_raw.replace(',', '.')
                dur_mins = int(float(dur_raw))
            except:
                dur_mins = 0
            
            if not tl_name:
                continue
                
            tl = Tasklist.query.filter_by(name=tl_name).first()
            if not tl:
                tl = Tasklist(name=tl_name, description=tl_name, project_code=p_code if p_code else None)
                db.session.add(tl)
                db.session.flush()
            else:
                if p_code and tl.project_code != p_code:
                    tl.project_code = p_code
                
            if proc:
                existing = TasklistProcedure.query.filter_by(tasklist_id=tl.id, name=proc).first()
                if not existing:
                    new_proc = TasklistProcedure(tasklist_id=tl.id, name=proc, requires_attachment=req_att, estimated_minutes=dur_mins)
                    db.session.add(new_proc)
                    counter += 1
        db.session.commit()
        flash(f'Tasklists imported easily. Added {counter} new procedures.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing Excel: {str(e)}', 'error')
        
    return redirect(url_for('settings.index'))

import ast

@settings_bp.route('/chiller_fault/import_dict', methods=['POST'])
@login_required
def import_chiller_fault_dict():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    chiller_type = request.form.get('chiller_type')
    category = request.form.get('category', 'general')
    dict_text = request.form.get('dict_text')
    
    if not chiller_type or not dict_text:
        flash('Chiller type and dictionary text are required.', 'error')
        return redirect(url_for('settings.index') + '#chiller-faults')
        
    try:
        parsed_dict = ast.literal_eval(dict_text.strip())
        
        if not isinstance(parsed_dict, dict):
            raise ValueError("Parsed data is not a valid Python dictionary")
            
        ChillerFaultCode.query.filter_by(chiller_type=chiller_type, category=category).delete()
        
        count = 0
        for code, desc in parsed_dict.items():
            new_fault = ChillerFaultCode(
                chiller_type=chiller_type,
                category=category,
                fault_code=str(code),
                description=str(desc)
            )
            db.session.add(new_fault)
            count += 1
            
        db.session.commit()
        flash(f'Successfully imported {count} fault codes for {chiller_type}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error parsing dictionary: {str(e)}', 'error')
        
    return redirect(url_for('settings.index') + '#chiller-faults')

@settings_bp.route('/chiller_fault/<int:id>/delete', methods=['POST'])
@login_required
def delete_chiller_fault(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    fault = ChillerFaultCode.query.get_or_404(id)
    db.session.delete(fault)
    db.session.commit()
    flash('Fault code deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#chiller-faults')


@settings_bp.route('/chiller_fault/delete_chiller/<string:chiller_type>', methods=['POST'])
@login_required
def delete_chiller_faults_by_type(chiller_type):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    ChillerFaultCode.query.filter_by(chiller_type=chiller_type).delete()
    db.session.commit()
    flash(f'All fault codes for {chiller_type} deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#chiller-faults')


@settings_bp.route('/chiller_fault/delete_category/<string:chiller_type>/<string:category>', methods=['POST'])
@login_required
def delete_chiller_faults_by_category(chiller_type, category):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    ChillerFaultCode.query.filter_by(chiller_type=chiller_type, category=category).delete()
    db.session.commit()
    flash(f'All {category} fault codes for {chiller_type} deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#chiller-faults')


# Helpdesk Master Data Routes

@settings_bp.route('/helpdesk_module/add', methods=['POST'])
@login_required
def add_hd_module():
    name = request.form.get('name')
    site_id = request.form.get('site_id')
    if name:
        new_mod = HelpdeskModule(name=name, site_id=site_id or None)
        db.session.add(new_mod)
        db.session.commit()
        flash('Helpdesk module added.', 'success')
    return redirect(url_for('settings.index') + '#hd-modules')

@settings_bp.route('/helpdesk_module/<int:id>/edit', methods=['POST'])
@login_required
def edit_hd_module(id):
    mod = HelpdeskModule.query.get_or_404(id)
    mod.name = request.form.get('name')
    mod.site_id = request.form.get('site_id') or None
    db.session.commit()
    flash('Helpdesk module updated.', 'success')
    return redirect(url_for('settings.index') + '#hd-modules')

@settings_bp.route('/helpdesk_module/<int:id>/delete', methods=['POST'])
@login_required
def delete_hd_module(id):
    mod = HelpdeskModule.query.get_or_404(id)
    db.session.delete(mod)
    db.session.commit()
    flash('Helpdesk module deleted.', 'success')
    return redirect(url_for('settings.index') + '#hd-modules')

@settings_bp.route('/helpdesk_location/add', methods=['POST'])
@login_required
def add_hd_location():
    name = request.form.get('name')
    site_id = request.form.get('site_id')
    if name:
        new_loc = HelpdeskLocation(name=name, site_id=site_id or None)
        db.session.add(new_loc)
        db.session.commit()
        flash('Helpdesk location added.', 'success')
    return redirect(url_for('settings.index') + '#hd-locations')

@settings_bp.route('/helpdesk_location/<int:id>/edit', methods=['POST'])
@login_required
def edit_hd_location(id):
    loc = HelpdeskLocation.query.get_or_404(id)
    loc.name = request.form.get('name')
    loc.site_id = request.form.get('site_id') or None
    db.session.commit()
    flash('Helpdesk location updated.', 'success')
    return redirect(url_for('settings.index') + '#hd-locations')

@settings_bp.route('/helpdesk_location/<int:id>/delete', methods=['POST'])
@login_required
def delete_hd_location(id):
    loc = HelpdeskLocation.query.get_or_404(id)
    db.session.delete(loc)
    db.session.commit()
    flash('Helpdesk location deleted.', 'success')
    return redirect(url_for('settings.index') + '#hd-locations')

# Helpdesk CSV Import/Export

@settings_bp.route('/helpdesk_module/template', methods=['GET'])
@login_required
def download_hd_module_template():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    import pandas as pd
    import io
    from flask import send_file
    df = pd.DataFrame(columns=['Name', 'Site ID'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, download_name="helpdesk_modules_template.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@settings_bp.route('/helpdesk_module/export', methods=['GET'])
@login_required
def export_hd_modules():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    import pandas as pd
    import io
    from flask import send_file
    data = []
    modules = HelpdeskModule.query.all()
    for m in modules:
        data.append({'Name': m.name, 'Site ID': m.site_id or ''})
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, download_name="helpdesk_modules.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@settings_bp.route('/helpdesk_module/import', methods=['POST'])
@login_required
def import_hd_modules():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    import pandas as pd
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file.', 'error')
        return redirect(url_for('settings.index') + '#hd-modules')
    try:
        df = pd.read_excel(file)
        count = 0
        for index, row in df.iterrows():
            name = str(row.get('Name', '')).strip()
            if name == 'nan' or not name: continue
            site_id = str(row.get('Site ID', '')).strip()
            site_id = int(float(site_id)) if site_id != 'nan' and site_id else None
            existing = HelpdeskModule.query.filter_by(name=name, site_id=site_id).first()
            if not existing:
                new_mod = HelpdeskModule(name=name, site_id=site_id)
                db.session.add(new_mod)
                count += 1
        db.session.commit()
        flash(f'Imported {count} helpdesk modules.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing Excel: {str(e)}', 'error')
    return redirect(url_for('settings.index') + '#hd-modules')
        
    try:
        decoded = file.stream.read().decode("utf-8-sig")
        stream = io.StringIO(decoded)
        csv_input = csv.DictReader(stream)
        
        count = 0
        for row in csv_input:
            name = row.get('Module Name', '').strip()
            site_name = row.get('Site Name', '').strip()
            
            if not name: continue
            
            site = Site.query.filter_by(name=site_name).first()
            if not site and site_name:
                continue # Skip if site specified but not found
                
            # Check if exists for this site
            existing = HelpdeskModule.query.filter_by(name=name, site_id=site.id if site else None).first()
            if not existing:
                new_mod = HelpdeskModule(name=name, site_id=site.id if site else None)
                db.session.add(new_mod)
                count += 1
                
        db.session.commit()
        flash(f'Imported {count} helpdesk modules.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing CSV: {str(e)}', 'error')
        
    return redirect(url_for('settings.index') + '#hd-modules')

@settings_bp.route('/helpdesk_location/template', methods=['GET'])
@login_required
def download_hd_location_template():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    import pandas as pd
    import io
    from flask import send_file
    df = pd.DataFrame(columns=['Name', 'Site ID'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, download_name="helpdesk_locations_template.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@settings_bp.route('/helpdesk_location/export', methods=['GET'])
@login_required
def export_hd_locations():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    import pandas as pd
    import io
    from flask import send_file
    data = []
    locations = HelpdeskLocation.query.all()
    for m in locations:
        data.append({'Name': m.name, 'Site ID': m.site_id or ''})
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, download_name="helpdesk_locations.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@settings_bp.route('/helpdesk_location/import', methods=['POST'])
@login_required
def import_hd_locations():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    import pandas as pd
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file.', 'error')
        return redirect(url_for('settings.index') + '#hd-locations')
    try:
        df = pd.read_excel(file)
        count = 0
        for index, row in df.iterrows():
            name = str(row.get('Name', '')).strip()
            if name == 'nan' or not name: continue
            site_id = str(row.get('Site ID', '')).strip()
            site_id = int(float(site_id)) if site_id != 'nan' and site_id else None
            existing = HelpdeskLocation.query.filter_by(name=name, site_id=site_id).first()
            if not existing:
                new_mod = HelpdeskLocation(name=name, site_id=site_id)
                db.session.add(new_mod)
                count += 1
        db.session.commit()
        flash(f'Imported {count} helpdesk locations.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing Excel: {str(e)}', 'error')
    return redirect(url_for('settings.index') + '#hd-locations')
        
    try:
        decoded = file.stream.read().decode("utf-8-sig")
        stream = io.StringIO(decoded)
        csv_input = csv.DictReader(stream)
        
        count = 0
        for row in csv_input:
            name = row.get('Location Name', '').strip()
            site_name = row.get('Site Name', '').strip()
            
            if not name: continue
            
            site = Site.query.filter_by(name=site_name).first()
            if not site and site_name:
                continue
                
            existing = HelpdeskLocation.query.filter_by(name=name, site_id=site.id if site else None).first()
            if not existing:
                new_loc = HelpdeskLocation(name=name, site_id=site.id if site else None)
                db.session.add(new_loc)
                count += 1
                
        db.session.commit()
        flash(f'Imported {count} helpdesk locations.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing CSV: {str(e)}', 'error')
        
    return redirect(url_for('settings.index') + '#hd-locations')

def format_excel_sheet(worksheet):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='BFBFBF'), 
                    right=Side(style='thin', color='BFBFBF'),
                    top=Side(style='thin', color='BFBFBF'), 
                    bottom=Side(style='thin', color='BFBFBF'))

    # Format header row
    for col_num in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Auto-adjust column widths and add borders
    for col in worksheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            
            # Apply border to all cells
            cell.border = border
            
            # Formatting for Price and Cost columns (D and E usually)
            if cell.row > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                
        adjusted_width = max(12, min(max_length + 2, 50))
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

@settings_bp.route('/part/template', methods=['GET'])
@login_required
def download_part_template():
    import pandas as pd
    import io
    from flask import send_file
    from openpyxl.worksheet.datavalidation import DataValidation
    
    df = pd.DataFrame(columns=[
        'Product Code', 'Name', 'Category', 'Unit', 'Price', 'Unit Cost', 'Currency'
    ])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Master Parts')
        worksheet = writer.sheets['Master Parts']
        
        # Add Data Validation for Currency (Column G, rows 2 to 1000)
        dv = DataValidation(type="list", formula1='"IDR,USD"', allow_blank=True)
        dv.add('G2:G1000')
        worksheet.add_data_validation(dv)
        
        # Format the Excel sheet nicely
        format_excel_sheet(worksheet)
        
    output.seek(0)
    response = send_file(output, download_name='master_parts_template.xlsx', as_attachment=True)
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@settings_bp.route('/part/export', methods=['GET'])
@login_required
def export_parts():
    import pandas as pd
    import io
    from flask import send_file
    
    parts = Part.query.all()
    data = []
    for p in parts:
        data.append({
            'Product Code': p.code,
            'Name': p.name,
            'Category': p.category.name if p.category else '',
            'Unit': p.unit or '',
            'Price': p.price,
            'Unit Cost': p.unit_cost,
            'Currency': p.currency
        })
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Master Parts')
        worksheet = writer.sheets['Master Parts']
        
        # Format the Excel sheet nicely
        format_excel_sheet(worksheet)
    output.seek(0)
    return send_file(output, download_name='master_parts.xlsx', as_attachment=True)

@settings_bp.route('/part/import', methods=['POST'])
@login_required
def import_parts():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('settings.index') + '#parts')
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('settings.index') + '#parts')
    if file and file.filename.endswith(('.xls', '.xlsx')):
        try:
            import pandas as pd
            df = pd.read_excel(file)
            count = 0
            for index, row in df.iterrows():
                code = str(row.get('Product Code', '')).strip()
                name = str(row.get('Name', '')).strip()
                if not code or not name or code == 'nan' or name == 'nan':
                    continue
                
                category_name = str(row.get('Category', '')).strip()
                unit = str(row.get('Unit', '')).strip()
                price = row.get('Price', 0.0)
                unit_cost = row.get('Unit Cost', 0.0)
                currency = str(row.get('Currency', 'IDR')).strip()
                
                if pd.isna(price): price = 0.0
                if pd.isna(unit_cost): unit_cost = 0.0
                if pd.isna(unit) or unit == 'nan': unit = None
                if pd.isna(currency) or currency == 'nan': currency = 'IDR'
                
                category_id = None
                if category_name and category_name != 'nan':
                    cat = Category.query.filter_by(name=category_name, type='Part').first()
                    if not cat:
                        cat = Category(name=category_name, type='Part')
                        db.session.add(cat)
                        db.session.commit()
                    category_id = cat.id
                
                part = Part.query.filter_by(code=code).first()
                if not part:
                    part = Part(code=code)
                    db.session.add(part)
                
                part.name = name
                part.unit = unit
                part.price = float(price)
                part.unit_cost = float(unit_cost)
                part.currency = currency
                part.category_id = category_id
                
                count += 1
            
            db.session.commit()
            flash(f'Successfully imported {count} master parts.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing file: {str(e)}', 'error')
    else:
        flash('Invalid file format. Please upload an Excel file.', 'error')
    
    return redirect(url_for('settings.index') + '#parts')

@settings_bp.route('/part/create', methods=['GET', 'POST'])
@login_required
def create_part():
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        unit = request.form.get('unit')
        price = request.form.get('price', 0.0)
        unit_cost = request.form.get('unit_cost', 0.0)
        currency = request.form.get('currency', 'IDR')
        category_name = request.form.get('category_name')
        
        category_id = None
        if category_name:
            category_name = category_name.strip()
            cat = Category.query.filter_by(name=category_name, type='Part').first()
            if not cat:
                cat = Category(name=category_name, type='Part')
                db.session.add(cat)
                db.session.commit()
            category_id = cat.id
        # basic generation if code is empty
        if not code:
            import random
            code = f"PRT{random.randint(1000, 9999)}"
            
        new_part = Part(
            name=name, code=code, unit=unit,
            price=float(price), unit_cost=float(unit_cost),
            currency=currency,
            site_id=None, # Master parts don't necessarily need site_id but we can leave it null
            category_id=category_id if category_id else None
        )
        db.session.add(new_part)
        db.session.commit()
        flash('Master Part created successfully!', 'success')
        return redirect(url_for('settings.index') + '#parts')
        
    categories = Category.query.filter_by(type='Part').all()
    return render_template('settings/create_part.html', categories=categories)

@settings_bp.route('/part/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_part(id):
    part = Part.query.get_or_404(id)
    if request.method == 'POST':
        part.name = request.form.get('name')
        part.code = request.form.get('code')
        part.unit = request.form.get('unit')
        part.price = float(request.form.get('price', 0.0))
        part.unit_cost = float(request.form.get('unit_cost', 0.0))
        part.currency = request.form.get('currency', 'IDR')
        
        category_name = request.form.get('category_name')
        if category_name:
            category_name = category_name.strip()
            cat = Category.query.filter_by(name=category_name, type='Part').first()
            if not cat:
                cat = Category(name=category_name, type='Part')
                db.session.add(cat)
                db.session.commit()
            part.category_id = cat.id
        else:
            part.category_id = None
        
        db.session.commit()
        flash('Master Part updated successfully!', 'success')
        return redirect(url_for('settings.index') + '#parts')
        
    categories = Category.query.filter_by(type='Part').all()
    return render_template('settings/edit_part.html', part=part, categories=categories)

@settings_bp.route('/part/<int:id>/delete', methods=['POST'])
@login_required
def delete_part(id):
    if current_user.role != 'Admin':
        flash('Access denied.', 'danger')
        return redirect(url_for('settings.index') + '#parts')
        
    part = Part.query.get_or_404(id)
    try:
        # StockLevel will cascade or need to be handled, but assuming cascade is ok
        # Actually let's delete stock levels explicitly just in case if no cascade is set
        from models import StockTransaction, WorkOrderPart, AssetPartBOM, POItem
        StockTransaction.query.filter_by(part_id=id).delete()
        WorkOrderPart.query.filter_by(part_id=id).delete()
        AssetPartBOM.query.filter_by(part_id=id).delete()
        POItem.query.filter_by(part_id=id).delete()
        StockLevel.query.filter_by(part_id=id).delete()
        
        db.session.delete(part)
        db.session.commit()
        flash('Master Part deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting part: {str(e)}', 'error')
        
    return redirect(url_for('settings.index') + '#parts')

@settings_bp.route('/customer/add', methods=['POST'])
@login_required
def add_customer():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    new_customer = Customer(
        partner_organization=request.form.get('partner_organization'),
        point_of_contact=request.form.get('point_of_contact'),
        email=request.form.get('email'),
        mobile_number=request.form.get('mobile_number')
    )
    db.session.add(new_customer)
    db.session.commit()
    flash('Customer added successfully', 'success')
    return redirect(url_for('settings.index') + '#customers')

@settings_bp.route('/customer/edit/<int:id>', methods=['POST'])
@login_required
def edit_customer(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    customer = Customer.query.get_or_404(id)
    customer.partner_organization = request.form.get('partner_organization')
    customer.point_of_contact = request.form.get('point_of_contact')
    customer.email = request.form.get('email')
    customer.mobile_number = request.form.get('mobile_number')
    db.session.commit()
    flash('Customer updated successfully', 'success')
    return redirect(url_for('settings.index') + '#customers')

@settings_bp.route('/customer/delete/<int:id>', methods=['POST'])
@login_required
def delete_customer(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    customer = Customer.query.get_or_404(id)
    db.session.delete(customer)
    db.session.commit()
    flash('Customer deleted successfully', 'success')
    return redirect(url_for('settings.index') + '#customers')

@settings_bp.route('/logsheet_template/add', methods=['POST'])
@login_required
def add_logsheet_template():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    name = request.form.get('name')
    if name:
        existing = LogsheetTemplate.query.filter_by(name=name).first()
        if existing:
            flash(f'Logsheet template named {name} already exists.', 'error')
        else:
            new_lt = LogsheetTemplate(name=name)
            db.session.add(new_lt)
            db.session.commit()
            flash('Logsheet template added successfully.', 'success')
    return redirect(url_for('settings.index') + '#logsheet-templates')

@settings_bp.route('/logsheet_template/<int:id>/edit', methods=['POST'])
@login_required
def edit_logsheet_template(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    lt = LogsheetTemplate.query.get_or_404(id)
    name = request.form.get('name')
    
    if name:
        existing = LogsheetTemplate.query.filter(LogsheetTemplate.name == name, LogsheetTemplate.id != id).first()
        if existing:
            flash(f'A logsheet template named {name} already exists.', 'error')
        else:
            lt.name = name
            
            # Sync parameters if provided via JSON
            params_json_str = request.form.get('parameters_json')
            if params_json_str:
                import json
                try:
                    params_data = json.loads(params_json_str)
                    
                    # Delete existing parameters
                    LogsheetTemplateParameter.query.filter_by(template_id=lt.id).delete()
                    
                    # Add new parameters
                    for idx, p in enumerate(params_data):
                        new_p = LogsheetTemplateParameter(
                            template_id=lt.id,
                            name=p.get('name'),
                            entry_type=p.get('entry_type', 'reading'),
                            unit=p.get('unit'),
                            standard_min=p.get('min') if p.get('min') != '' else None,
                            standard_max=p.get('max') if p.get('max') != '' else None,
                            position=idx
                        )
                        db.session.add(new_p)
                except Exception as e:
                    flash(f'Error saving parameters: {str(e)}', 'error')
            
            db.session.commit()
            flash('Logsheet template updated successfully.', 'success')
            
    return redirect(url_for('settings.index') + '#logsheet-templates')

@settings_bp.route('/logsheet_template/<int:id>/delete', methods=['POST'])
@login_required
def delete_logsheet_template(id):
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    lt = LogsheetTemplate.query.get_or_404(id)
    LogsheetTemplateParameter.query.filter_by(template_id=lt.id).delete()
    db.session.delete(lt)
    db.session.commit()
    flash('Logsheet template deleted successfully.', 'success')
    return redirect(url_for('settings.index') + '#logsheet-templates')

@settings_bp.route('/api/logsheet_template/<int:id>/parameters')
def api_logsheet_template_parameters(id):
    lt = LogsheetTemplate.query.get_or_404(id)
    params = lt.parameters.all()
    data = []
    for p in params:
        data.append({
            'id': p.id,
            'name': p.name,
            'entry_type': p.entry_type,
            'unit': p.unit,
            'min': p.standard_min,
            'max': p.standard_max
        })
    from flask import jsonify
    return jsonify(data)

